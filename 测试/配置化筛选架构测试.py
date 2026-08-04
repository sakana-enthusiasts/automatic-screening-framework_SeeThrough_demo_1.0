from __future__ import annotations

from io import BytesIO
from importlib import import_module
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from 核心系统.数据管理接口 import 文件数据访问管理器
from 核心系统.流程控制器 import 创建颅骨透明化筛选流程控制器
from 核心系统.运行数据管理 import 运行数据管理器
from 核心系统.通用规则引擎 import 属性注册表, 毒性判定指标插件, 通用规则执行器, 规则注册表
from 插件.数据导入.通用候选表导入插件 import 通用候选表导入插件
from 插件.化学计算.补充数据2汉森距离计算插件 import 补充数据2汉森距离计算插件


项目根目录 = Path(__file__).resolve().parents[1]


def 完整映射() -> dict[str, str | None]:
    return {
        "化学名称列": "名称", "CAS列": "CAS", "货号列": "货号", "候选编号列": "编号",
        "水合能力平均值列": "水合", "水合能力标准差列": "标准差", "eRI列": "eRI",
        "dD列": "dD", "dP列": "dP", "dH列": "dH", "pH列": None,
        "实测RI列": None, "气味列": None, "毒性或安全性列": None, "实际互溶状态列": None,
    }


def 完整CSV() -> bytes:
    return "编号,名称,CAS,货号,水合,标准差,eRI,dD,dP,dH\nU1,=Formula,100-51-6,X1,-1.2,0.1,1.61,18,8,7\n".encode("utf-8")


def 完整XLSX() -> bytes:
    工作簿 = Workbook()
    工作表 = 工作簿.active
    工作表.append(["编号", "名称", "CAS", "货号", "水合", "标准差", "eRI", "dD", "dP", "dH"])
    工作表.append(["U2", "Benzyl alcohol", "100-51-6", "X2", -1.2, 0.1, 1.61, 18, 8, 7])
    缓冲区 = BytesIO()
    工作簿.save(缓冲区)
    return 缓冲区.getvalue()


def 运行用户流程(临时目录: Path, 文件名: str, 内容: bytes, 配置: str = "generic_compound") -> dict:
    控制器 = 创建颅骨透明化筛选流程控制器(临时目录)
    return 控制器.执行用户候选导入(
        文件名,
        内容,
        完整映射(),
        {"应用配置": 配置, "允许网络身份查询": False},
        应用配置=配置,
    )


def test_无标签端到端流程可完成(monkeypatch: pytest.MonkeyPatch) -> None:
    控制器 = 创建颅骨透明化筛选流程控制器(项目根目录)
    原读取 = 控制器.数据管理器.读取软件数据库表格

    def 无标签(文件名: str):
        if 文件名 == "补充数据2论文最终10标签表.csv":
            raise FileNotFoundError(文件名)
        return 原读取(文件名)

    monkeypatch.setattr(控制器.数据管理器, "读取软件数据库表格", 无标签)
    摘要 = 控制器.执行补充数据2规则筛选流程()
    assert 摘要["自动规则最终剩余数"] == 20


def test_有标签端到端追加验证而不改变自动候选() -> None:
    控制器 = 创建颅骨透明化筛选流程控制器(项目根目录)
    摘要 = 控制器.执行补充数据2规则筛选流程()
    验证 = 控制器.数据管理器.读取筛选结果("补充数据2_论文标签验证统计")
    assert 摘要["自动规则最终剩余数"] == 20
    assert 验证.loc[0, "标签验证状态"] == "已执行"


def test_完整CSV可端到端运行且报告转义公式(tmp_path: Path) -> None:
    摘要 = 运行用户流程(tmp_path, "complete.csv", 完整CSV())
    assert 摘要["导入记录数"] == 1
    报告 = Path(摘要["报告路径"])
    assert 报告.is_file() and 摘要["run_id"] in str(报告)
    工作簿 = load_workbook(报告, data_only=False)
    assert 工作簿["候选筛选结果"][2][0].value == "'=Formula"


def test_完整XLSX可端到端运行(tmp_path: Path) -> None:
    摘要 = 运行用户流程(tmp_path, "complete.xlsx", 完整XLSX())
    assert 摘要["导入记录数"] == 1 and Path(摘要["报告路径"]).is_file()


def test_缺失字段显示无法评估而不是零或排除(tmp_path: Path) -> None:
    控制器 = 创建颅骨透明化筛选流程控制器(tmp_path)
    摘要 = 控制器.执行用户候选导入(
        "missing.csv", "编号,名称,CAS\nU3,Only name,100-51-6\n".encode("utf-8"),
        {"化学名称列": "名称", "CAS列": "CAS", "货号列": None, "候选编号列": "编号"},
        {"应用配置": "seethrough_aqueous", "允许网络身份查询": False},
    )
    结果 = 控制器.数据管理器.读取筛选结果("用户导入_规则筛选结果")
    assert 摘要["无法评估数"] == 1
    assert 结果.loc[0, "规则总状态"] == "无法评估"
    assert "缺少属性" in 结果.loc[0, "无法执行规则原因"]


def test_非SeeThrough运行编号彼此隔离(tmp_path: Path) -> None:
    第一轮 = 运行用户流程(tmp_path, "a.csv", 完整CSV(), "generic_compound")
    第二轮 = 运行用户流程(tmp_path, "b.xlsx", 完整XLSX(), "user_custom")
    assert 第一轮["run_id"] != 第二轮["run_id"]
    根目录 = tmp_path / "数据" / "运行记录"
    assert (根目录 / 第一轮["run_id"] / "筛选结果" / "用户导入_规则筛选结果.csv").is_file()
    assert (根目录 / 第二轮["run_id"] / "筛选结果" / "用户导入_规则筛选结果.csv").is_file()


def test_路径穿越被拒绝(tmp_path: Path) -> None:
    管理器 = 运行数据管理器(tmp_path)
    with pytest.raises(ValueError, match="路径"):
        通用候选表导入插件().执行({
            "数据管理器": 管理器, "文件名": "../traversal.csv", "文件内容": "名称\nA\n".encode("utf-8"), "字段映射": {"化学名称列": "名称"},
        })


def test_重复列名被阻止进入流程(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="重复列名"):
        通用候选表导入插件().执行({
            "数据管理器": 文件数据访问管理器(tmp_path), "文件名": "duplicate.csv", "文件内容": "名称,名称\nA,B\n".encode("utf-8"), "字段映射": {"化学名称列": "名称"},
        })


def test_扩展名与实际文件签名必须一致(tmp_path: Path) -> None:
    管理器 = 文件数据访问管理器(tmp_path)
    with pytest.raises(ValueError, match="签名"):
        通用候选表导入插件().执行({
            "数据管理器": 管理器, "文件名": "fake.xlsx", "文件内容": b"name\nA\n", "字段映射": {"化学名称列": "name"},
        })


def test_HSP参照唯一且参数缺失显示无法评估(tmp_path: Path) -> None:
    管理器 = 运行数据管理器(tmp_path)
    输入 = pd.DataFrame([
        {"候选编号": "BA", "论文_CAS号": "100-51-6", "论文_dD": 18.0, "论文_dP": 8.0, "论文_dH": 7.0, "数值初筛通过": True},
        {"候选编号": "VA", "论文_CAS号": "93-03-8", "论文_dD": 17.0, "论文_dP": 7.0, "论文_dH": 6.0, "数值初筛通过": True},
        {"候选编号": "C1", "论文_CAS号": "1-11-1", "论文_dD": None, "论文_dP": 2.0, "论文_dH": 3.0, "数值初筛通过": True},
    ])
    管理器.保存中间结果("补充数据2_数值初筛记录", 输入)
    结果 = 补充数据2汉森距离计算插件().执行({"数据管理器": 管理器})
    assert set(结果.loc[结果["候选编号"].eq("C1"), "规则状态"]) == {"无法评估"}
    重复BA = pd.concat([输入, 输入.iloc[[0]]], ignore_index=True)
    管理器.保存中间结果("补充数据2_数值初筛记录", 重复BA)
    with pytest.raises(ValueError, match="必须唯一"):
        补充数据2汉森距离计算插件().执行({"数据管理器": 管理器})


def test_规则启用禁用决定跳过或无法评估() -> None:
    注册表 = 规则注册表.默认()
    执行器 = 通用规则执行器(注册表=属性注册表.默认())
    属性 = pd.DataFrame(columns=["run_id", "候选编号", "属性编号", "当前值"])
    全禁用 = 注册表.规则列表("seethrough_aqueous", {规则.规则编号: False for 规则 in 注册表.规则列表("seethrough_aqueous")})
    跳过 = 执行器.执行("r1", ["C1"], 属性, 全禁用)
    启用 = 注册表.规则列表("seethrough_aqueous", {"ST-AQ-001": True, "ST-AQ-002": False, "ST-AQ-003": False})
    无法评估 = 执行器.执行("r1", ["C1"], 属性, 启用)
    assert set(跳过["规则状态"]) == {"跳过"}
    assert "无法评估" in set(无法评估["规则状态"])


def test_毒性规则以独立证据指标接入规则引擎() -> None:
    证据 = pd.DataFrame([{
        "候选编号": "C1", "物种": "大鼠", "给药途径": "口服", "剂量": "10 mg/kg", "暴露时长": "24 h",
        "毒性终点": "Ames 遗传毒性阳性", "实验或预测": "实验", "来源": "文献", "可信度": "高",
    }])
    指标 = 毒性判定指标插件().执行("tox-run", 证据)
    属性 = 指标.rename(columns={"指标编号": "属性编号"})[["run_id", "候选编号", "属性编号", "当前值"]]
    规则 = 规则注册表.默认().规则列表("seethrough_aqueous", {"ST-AQ-001": False, "ST-AQ-002": False, "ST-AQ-003": False, "TOX-001": True})
    结果 = 通用规则执行器().执行("tox-run", ["C1"], 属性, 规则)
    assert "警告" in set(结果["规则状态"])


def test_四个未来页面显示各自接口状态(monkeypatch: pytest.MonkeyPatch) -> None:
    class 记录器:
        def __init__(self) -> None:
            self.标题: list[str] = []
            self.说明: list[str] = []

        def header(self, 文本: str) -> None:
            self.标题.append(文本)

        def info(self, 文本: str) -> None:
            self.说明.append(文本)

    页面 = [
        ("软件界面.配方构建页面", "渲染配方构建页面", "配方构建"),
        ("软件界面.透明化预测页面", "渲染透明化预测页面", "透明化预测"),
        ("软件界面.实验记录页面", "渲染实验记录页面", "实验记录"),
        ("软件界面.实验优化页面", "渲染实验优化页面", "实验优化"),
    ]
    for 模块名, 函数名, 预期标题 in 页面:
        模块 = import_module(模块名)
        记录 = 记录器()
        monkeypatch.setattr(模块, "st", 记录)
        getattr(模块, 函数名)()
        assert 记录.标题 == [预期标题] and 记录.说明
