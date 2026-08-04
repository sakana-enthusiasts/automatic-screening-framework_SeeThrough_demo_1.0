"""受限且可审计的 CSV/XLSX 用户候选表导入插件。"""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import PureWindowsPath
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from 插件.插件接口 import 基础插件接口


class 通用候选表导入插件(基础插件接口):
    插件标识 = "通用候选表导入"
    输出标识 = "用户导入_统一候选记录"
    最大上传字节数 = 2 * 1024 * 1024
    支持扩展名 = {".csv", ".xlsx"}

    字段映射 = {
        "化学名称列": "化学名称",
        "CAS列": "CAS号",
        "货号列": "货号",
        "候选编号列": "候选编号",
        "水合能力平均值列": "水合评分平均值",
        "水合能力标准差列": "水合评分标准差",
        "eRI列": "eRI",
        "pH列": "预测pH",
        "dD列": "dD",
        "dP列": "dP",
        "dH列": "dH",
        "实测RI列": "实测RI",
        "气味列": "气味",
        "毒性或安全性列": "毒性或安全性",
        "实际互溶状态列": "实际互溶状态",
    }

    @classmethod
    def 安全文件名(cls, 文件名: str) -> str:
        原始 = str(文件名 or "")
        标准化 = 原始.replace("\\", "/")
        if not 标准化 or "/" in 标准化 or ":" in 标准化 or ".." in 标准化.split("/"):
            raise ValueError("上传文件名必须为不含路径、盘符或父目录的单一文件名")
        名称 = PureWindowsPath(标准化).name
        if 名称 != 标准化:
            raise ValueError("上传文件名不安全")
        return 名称

    @classmethod
    def _检查扩展名与签名(cls, 文件名: str, 内容: bytes) -> str:
        安全名称 = cls.安全文件名(文件名)
        后缀 = PureWindowsPath(安全名称).suffix.lower()
        if 后缀 not in cls.支持扩展名:
            raise ValueError("仅支持 CSV 或 XLSX 文件")
        if not 内容:
            raise ValueError("上传文件为空")
        if len(内容) > cls.最大上传字节数:
            raise ValueError(f"上传文件超过大小上限：{cls.最大上传字节数} 字节")
        if 后缀 == ".xlsx" and not 内容.startswith(b"PK\x03\x04"):
            raise ValueError("XLSX 扩展名与实际文件签名不匹配")
        if 后缀 == ".csv":
            if 内容.startswith(b"PK\x03\x04") or b"\x00" in 内容[:4096]:
                raise ValueError("CSV 扩展名与实际文件内容不匹配")
            try:
                内容.decode("utf-8-sig")
            except UnicodeDecodeError as 错误:
                raise ValueError("CSV 必须使用 UTF-8 或 UTF-8-SIG 编码") from 错误
        return 后缀

    @staticmethod
    def _检查重复列(列名: list[Any]) -> None:
        标准列名 = [str(列).strip() for 列 in 列名]
        重复 = sorted({列 for 列 in 标准列名 if 标准列名.count(列) > 1})
        if 重复:
            raise ValueError(f"上传表含重复列名，不能进入流程：{重复}")
        if any(not 列 for 列 in 标准列名):
            raise ValueError("上传表存在空列名")

    @classmethod
    def 读取上传内容(cls, 文件名: str, 内容: bytes) -> pd.DataFrame:
        后缀 = cls._检查扩展名与签名(文件名, 内容)
        try:
            if 后缀 == ".csv":
                文本 = 内容.decode("utf-8-sig")
                读取器 = csv.reader(StringIO(文本))
                首行 = next(读取器, None)
                if 首行 is None:
                    raise ValueError("CSV 没有表头")
                cls._检查重复列(首行)
                表格 = pd.read_csv(StringIO(文本), dtype=object, keep_default_na=False)
            else:
                工作簿 = load_workbook(BytesIO(内容), read_only=True, data_only=True)
                工作表 = 工作簿.active
                首行 = [单元格.value for 单元格 in next(工作表.iter_rows(max_row=1), ())]
                cls._检查重复列(首行)
                表格 = pd.read_excel(BytesIO(内容), dtype=object, keep_default_na=False)
        except (ValueError, OSError, pd.errors.ParserError) as 错误:
            raise ValueError(f"无法读取上传文件：{错误}") from 错误
        if 表格.empty:
            raise ValueError("上传表没有数据行")
        表格.columns = [str(列).strip() for 列 in 表格.columns]
        return 表格

    @staticmethod
    def _映射列(原始表: pd.DataFrame, 映射: dict[str, str | None], 映射键: str, 输出列: str) -> pd.Series:
        来源列 = 映射.get(映射键)
        if not 来源列:
            return pd.Series(pd.NA, index=原始表.index, dtype="object")
        if 来源列 not in 原始表.columns:
            raise ValueError(f"字段映射不存在：{映射键} -> {来源列}")
        值 = 原始表[来源列].replace("", pd.NA)
        return 值.astype("object")

    def 执行(self, 数据上下文: dict[str, Any]) -> pd.DataFrame:
        数据管理器 = 数据上下文["数据管理器"]
        文件名 = self.安全文件名(str(数据上下文["文件名"]))
        内容 = bytes(数据上下文["文件内容"])
        原始表 = self.读取上传内容(文件名, 内容)
        映射 = dict(数据上下文.get("字段映射") or {})
        名称列 = 映射.get("化学名称列")
        if not 名称列:
            raise ValueError("必须映射化学名称列")
        if 名称列 not in 原始表.columns:
            raise ValueError(f"化学名称列不存在：{名称列}")
        if hasattr(数据管理器, "当前运行编号") and not 数据管理器.当前运行编号:
            数据管理器.创建运行(str(数据上下文.get("应用配置", "user_custom")), 文件名)
        run_id = getattr(数据管理器, "当前运行编号", "legacy") or "legacy"
        数据管理器.保存用户导入原始文件(文件名, 内容)
        统一 = pd.DataFrame(index=原始表.index)
        for 映射键, 输出列 in self.字段映射.items():
            统一[输出列] = self._映射列(原始表, 映射, 映射键, 输出列)
            if 映射键 not in {"化学名称列", "CAS列", "货号列", "候选编号列"}:
                来源列 = 映射.get(映射键)
                统一[f"字段状态_{输出列}"] = "用户表格提供" if 来源列 else "缺失"
        自动编号 = [f"U{序号:05d}" for 序号 in range(1, len(统一) + 1)]
        统一["候选编号"] = 统一["候选编号"].fillna(pd.Series(自动编号, index=统一.index)).astype(str).str.strip()
        统一["化学名称"] = 统一["化学名称"].astype(str).str.strip()
        if 统一["化学名称"].eq("").any():
            raise ValueError("化学名称不能为空")
        统一["CAS号"] = 统一["CAS号"].fillna("").astype(str).str.strip()
        统一["货号"] = 统一["货号"].fillna("").astype(str).str.strip()
        统一["run_id"] = run_id
        统一["候选类型"] = str(数据上下文.get("候选类型", "通用候选"))
        统一["来源文件名"] = 文件名
        统一["导入状态"] = "可用"
        统一["导入冲突说明"] = ""
        重复编号 = 统一["候选编号"].duplicated(keep=False)
        统一.loc[重复编号, "导入状态"] = "重复编号"
        统一.loc[重复编号, "导入冲突说明"] = "重复编号"
        有效CAS = 统一["CAS号"].ne("")
        重复CAS = 有效CAS & 统一["CAS号"].duplicated(keep=False)
        统一.loc[重复CAS, "导入冲突说明"] = 统一.loc[重复CAS, "导入冲突说明"].replace("", "重复CAS")
        CAS格式异常 = 有效CAS & ~统一["CAS号"].str.fullmatch(r"\d{2,7}-\d{2}-\d", na=False)
        统一.loc[CAS格式异常, "导入冲突说明"] = 统一.loc[CAS格式异常, "导入冲突说明"].replace("", "CAS格式异常")
        同名多CAS = 统一.groupby("化学名称")["CAS号"].transform(lambda 值: 值[值.ne("")].nunique() > 1)
        统一.loc[同名多CAS, "导入冲突说明"] = 统一.loc[同名多CAS, "导入冲突说明"].replace("", "同名不同CAS")
        数据管理器.保存中间结果(self.输出标识, 统一)
        return 统一
