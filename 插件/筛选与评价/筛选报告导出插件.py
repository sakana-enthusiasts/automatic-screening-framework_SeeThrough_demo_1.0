"""导出 SeeThrough 配置运行的动态审计报告，不使用固定数量。"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import subprocess
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rdkit import rdBase

from 插件.插件接口 import 基础插件接口


class 筛选报告导出插件(基础插件接口):
    插件标识 = "筛选报告导出"
    文件名 = "SeeThrough规则筛选演示报告.xlsx"

    @staticmethod
    def _读取(数据管理器: Any, 方法: str, 标识: str) -> pd.DataFrame:
        try:
            return getattr(数据管理器, 方法)(标识)
        except FileNotFoundError:
            return pd.DataFrame()

    @staticmethod
    def _安全单元格(值: Any) -> Any:
        if isinstance(值, (list, tuple, set, dict)):
            值 = "|".join(map(str, 值)) if not isinstance(值, dict) else str(值)
        if isinstance(值, str) and 值.lstrip().startswith(("=", "+", "-", "@")):
            return "'" + 值
        return None if pd.isna(值) else 值

    @classmethod
    def _写入工作表(cls, 工作簿: Workbook, 名称: str, 数据: pd.DataFrame) -> None:
        工作表 = 工作簿.create_sheet(名称)
        数据 = 数据.copy()
        if 数据.empty and len(数据.columns) == 0:
            数据 = pd.DataFrame({"说明": ["当前无记录"]})
        工作表.append([cls._安全单元格(列) for 列 in 数据.columns])
        for 行 in 数据.itertuples(index=False, name=None):
            工作表.append([cls._安全单元格(值) for 值 in 行])
        工作表.freeze_panes = "A2"
        工作表.auto_filter.ref = 工作表.dimensions
        for 单元格 in 工作表[1]:
            单元格.font = Font(bold=True, color="FFFFFF")
            单元格.fill = PatternFill("solid", fgColor="176B87")
            单元格.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for 行 in 工作表.iter_rows(min_row=2):
            for 单元格 in 行:
                单元格.alignment = Alignment(vertical="top", wrap_text=True)
        for 列号, 列名 in enumerate(数据.columns, 1):
            最大宽度 = max([len(str(列名)) + 2, *[len(str(值)) + 1 for 值 in 数据.iloc[:100, 列号 - 1].fillna("")]])
            工作表.column_dimensions[get_column_letter(列号)].width = min(max(10, 最大宽度), 42)

    def 执行(self, 数据上下文: dict[str, Any]) -> Path:
        数据管理器 = 数据上下文["数据管理器"]
        记录 = 数据管理器.读取筛选结果("补充数据2_规则筛选统一记录")
        自动掩码 = 记录.get("自动规则通过", pd.Series(False, index=记录.index)).astype(str).str.lower().eq("true")
        自动 = 记录.loc[自动掩码].copy()
        标签掩码 = 自动.get("论文最终10候选标签", pd.Series("未提供验证标签", index=自动.index)).eq("是")
        标签候选 = 自动.loc[标签掩码].copy()
        导入审计 = self._读取(数据管理器, "读取中间结果", "补充数据2_导入清理审计")
        步骤 = self._读取(数据管理器, "读取中间结果", "补充数据2_规则筛选步骤统计")
        结构 = self._读取(数据管理器, "读取中间结果", "补充数据2_结构映射结果")
        描述符 = self._读取(数据管理器, "读取中间结果", "补充数据2_RDKit描述符结果")
        距离 = self._读取(数据管理器, "读取中间结果", "补充数据2_汉森距离结果")
        身份 = self._读取(数据管理器, "读取中间结果", "补充数据2_41候选身份映射")
        身份冲突 = self._读取(数据管理器, "读取筛选结果", "化合物身份冲突表")
        冲突 = self._读取(数据管理器, "读取筛选结果", "补充数据2_官方数据冲突记录")
        PubChem = self._读取(数据管理器, "读取筛选结果", "PubChem毒性证据")
        动物模型 = self._读取(数据管理器, "读取筛选结果", "当前动物实验模型结果")
        人体模型 = self._读取(数据管理器, "读取筛选结果", "未来人体口服模型结果")
        基准物 = self._读取(数据管理器, "读取筛选结果", "项目基准物清单")
        数据缺口 = 动物模型[[列 for 列 in ("候选编号", "化学名称", "数据不足项") if 列 in 动物模型.columns]].copy() if not 动物模型.empty else pd.DataFrame()
        验证 = self._读取(数据管理器, "读取筛选结果", "补充数据2_论文标签验证统计")
        try:
            提交 = subprocess.run(["git", "rev-parse", "--short", "HEAD"], capture_output=True, text=True, check=True).stdout.strip()
        except Exception:
            提交 = "当前目录不是Git仓库"
        原始行数 = "未记录"
        if not 导入审计.empty:
            候选列 = next((列 for 列 in 导入审计.columns if "原始" in 列 and "行" in 列), None)
            if 候选列:
                原始行数 = 导入审计.iloc[0][候选列]
        摘要 = pd.DataFrame({"项目": [
            "原始行数", "筛选统一记录数", "自动候选数", "论文标签重叠数", "结构成功数", "结构失败数", "数据冲突数", "运行时间", "软件版本或Git提交",
        ], "值": [
            原始行数, len(记录), len(自动), len(标签候选), int(结构.get("结构状态", pd.Series(dtype=str)).eq("已获得").sum()),
            int(len(结构) - structure_success(结构)), len(冲突), datetime.now(timezone.utc).isoformat(), f"RDKit {rdBase.rdkitVersion}; Git {提交}",
        ]})
        数据来源 = pd.DataFrame({"数据对象": ["候选与筛选结果", "论文标签", "结构查询", "Hansen参数", "毒性证据"], "来源与版本": [
            "SeeThrough Supplementary Data 2 / 当前运行记录", "Supplementary Data 3；仅可选验证", "PubChem PUG REST 或缓存", "Supplementary Data 2 / Hansen 距离公式", "PubChem缓存与独立证据记录",
        ]})
        工作簿 = Workbook()
        工作簿.remove(工作簿.active)
        for 名称, 数据 in (
            ("运行摘要", 摘要), ("筛选步骤", 步骤), ("自动候选", 自动), ("论文标签对照", 标签候选),
            ("论文标签验证", 验证), ("化合物身份映射", 身份), ("身份冲突", 身份冲突), ("结构映射", 结构),
            ("RDKit描述符", 描述符), ("Hansen距离", 距离), ("毒性原始证据", PubChem), ("动物实验模式结果", 动物模型),
            ("人体口服模式结果", 人体模型), ("项目基准物比较", 基准物), ("数据缺口", 数据缺口),
            ("官方数据冲突", 冲突), ("筛选统一记录", 记录), ("数据来源与版本", 数据来源),
        ):
            self._写入工作表(工作簿, 名称, 数据)
        缓冲区 = BytesIO()
        工作簿.save(缓冲区)
        return 数据管理器.保存导出报告(self.文件名, 缓冲区.getvalue())


def structure_success(结构: pd.DataFrame) -> int:
    return int(结构.get("结构状态", pd.Series(dtype=str)).eq("已获得").sum())
