"""按 run_id 导出通用候选筛选报告，并防止 Excel 公式注入。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from 核心系统.通用规则引擎 import 属性注册表, 规则注册表
from 插件.插件接口 import 基础插件接口


class 通用运行报告导出插件(基础插件接口):
    插件标识 = "通用运行报告导出"

    @staticmethod
    def _安全单元格(值: Any) -> Any:
        if isinstance(值, (list, tuple, set, dict)):
            值 = "|".join(map(str, 值)) if not isinstance(值, dict) else str(值)
        if isinstance(值, str) and 值.lstrip().startswith(("=", "+", "-", "@")):
            return "'" + 值
        if pd.isna(值):
            return None
        return 值

    @classmethod
    def _写入工作表(cls, 工作簿: Workbook, 名称: str, 表格: pd.DataFrame) -> None:
        工作表 = 工作簿.create_sheet(名称)
        数据 = 表格.copy()
        if 数据.empty and not len(数据.columns):
            数据 = pd.DataFrame({"说明": ["当前无记录"]})
        工作表.append([cls._安全单元格(列) for 列 in 数据.columns])
        for 行 in 数据.itertuples(index=False, name=None):
            工作表.append([cls._安全单元格(值) for 值 in 行])
        工作表.freeze_panes = "A2"
        工作表.auto_filter.ref = 工作表.dimensions
        for 单元格 in 工作表[1]:
            单元格.font = Font(bold=True, color="FFFFFF")
            单元格.fill = PatternFill("solid", fgColor="176B87")
            单元格.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for 行 in 工作表.iter_rows(min_row=2):
            for 单元格 in 行:
                单元格.alignment = Alignment(vertical="top", wrap_text=True)
        for 列号, 列名 in enumerate(数据.columns, 1):
            长度 = max([len(str(列名)) + 2, *[len(str(值)) + 1 for 值 in 数据.iloc[:100, 列号 - 1].fillna("")]])
            工作表.column_dimensions[get_column_letter(列号)].width = min(max(12, 长度), 42)

    @staticmethod
    def _读取(数据管理器: Any, 方法: str, 标识: str) -> pd.DataFrame:
        try:
            return getattr(数据管理器, 方法)(标识)
        except FileNotFoundError:
            return pd.DataFrame()

    def 执行(self, 数据上下文: dict[str, Any]):
        数据管理器 = 数据上下文["数据管理器"]
        run_id = str(数据上下文.get("运行编号") or 数据管理器.当前运行编号)
        应用配置 = str(数据上下文.get("应用配置", "user_custom"))
        if not run_id or run_id == "None":
            raise ValueError("通用运行报告需要有效 run_id")
        候选 = self._读取(数据管理器, "读取筛选结果", "用户导入_规则筛选结果")
        规则 = pd.DataFrame()
        if hasattr(数据管理器, "运行目录"):
            try:
                规则 = 数据管理器._读取表格(数据管理器.运行目录() / "规则结果", "规则执行记录")
            except FileNotFoundError:
                pass
        属性 = pd.DataFrame()
        if hasattr(数据管理器, "读取属性记录"):
            try:
                属性 = 数据管理器.读取属性记录()
            except FileNotFoundError:
                pass
        毒性指标 = self._读取(数据管理器, "读取毒性结果", "毒性判定指标") if hasattr(数据管理器, "读取毒性结果") else pd.DataFrame()
        状态统计 = 候选.get("规则总状态", pd.Series(dtype=str)).value_counts(dropna=False).rename_axis("规则总状态").reset_index(name="候选数")
        摘要 = pd.DataFrame([
            {"项目": "run_id", "值": run_id},
            {"项目": "候选数", "值": len(候选)},
            {"项目": "规则记录数", "值": len(规则)},
            {"项目": "属性记录数", "值": len(属性)},
            {"项目": "毒性指标记录数", "值": len(毒性指标)},
        ])
        工作簿 = Workbook()
        工作簿.remove(工作簿.active)
        for 名称, 表格 in (
            ("运行摘要", 摘要),
            ("状态统计", 状态统计),
            ("候选筛选结果", 候选),
            ("规则执行记录", 规则),
            ("属性注册表", 属性注册表.默认().导出()),
            ("属性记录", 属性),
            ("毒性判定指标", 毒性指标),
            ("应用配置", 规则注册表.默认().导出规则(应用配置)),
        ):
            self._写入工作表(工作簿, 名称, 表格)
        缓冲区 = BytesIO()
        工作簿.save(缓冲区)
        return 数据管理器.保存导出报告(f"{run_id}_筛选报告.xlsx", 缓冲区.getvalue())
